from tkinter import *
import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from functools import partial
root=Tk()
root.geometry("500x400")

#f=Label(text="my label",bg="red",relief="raised",borderwidth=15,fg="gray",font=("courier",25,"bold"))
#f.pack()
#logo=PhotoImage(file="yazan.gif")

"""
lb=Button(root,text="my text",bg="red")
lb.pack(padx=20,pady=30)


lb=Button(root,text="my text",bg="green")
lb.place(x=50,y=30,width=200,height=100)
lb=Button(root,text="my text",bg="green")
lb.place(x=50,y=20,width=200,height=100)
#g.configure(text="changed",bg="blue")
"""
#number1=input()
data= pd.read_excel("books.xlsx")
data.to_excel("books.xlsx",index=False)
wb=Workbook()
wb=load_workbook('books.xlsx')
ws=wb.active
column_1=ws['A']
#print(column_1)
#print(data)
#column_1=ws['B']
def get_a():
    for cell in column_1:
        print(column_1)
        
#def get_b():
#    for cell in column_2:
#        pass


#menubar
menubar = Menu(root)  
file = Menu(menubar)  
file.add_command(label="New")
file.add_command(label="Save")    
file.add_command(label="Exit", command=root.quit)  
menubar.add_cascade(label="File", menu=file)  

edit = Menu(menubar)   
edit.add_command(label="Copy")  
edit.add_command(label="Paste")  
menubar.add_cascade(label="Edit", menu=edit)  
   
root.config(menu=menubar)

######################

#labels

A=Label(text="Explore Books",font=("Arial",25))
A.grid(row=0,column=1)

B1=Label(text="Enter Book Id")
B1.grid(row=1,sticky=W)
B2=Label(text="Book Title")
B2.grid(row=2,sticky=W)
B3=Label(text="Book Author")
B3.grid(row=3,sticky=W)
B4=Label(text="No. of Copies")
B4.grid(row=4,sticky=W)
B5=Label(text="Book Price")
B5.grid(row=5,sticky=W)

#entry

user1=Entry(root,width=3)
user1.grid(row=1,column=1,sticky=W)

#user=Entry(root)
#user.grid(row=1,column=1)

#buttons

lb1=Button(root,text="Search",relief=RAISED)
lb1.grid(row=1,column=1,sticky=W,padx=30,ipadx=20)
lb2=Button(root,text="View all",relief=RAISED,command=get_a)
lb2.grid(row=1,column=1,sticky=W,padx=170,ipadx=20)

#text

T=Text(root)
T.grid(row=6,column=1,sticky=W)
root.mainloop()